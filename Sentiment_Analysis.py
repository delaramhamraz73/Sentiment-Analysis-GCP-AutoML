from google.cloud import automl_v1
import os
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from google.cloud import storage
import datetime
import re

# ONLY THE SHEET NAME HAS TO CHANGE EVERYDAY BASED ON THE CURRENT DATA

project_id = '536714825571'
model_id = 'TST2829388664918769664'

# name of the excel file
myExcel = Path('May-2020-new.xlsx')

# name of the sheet
sheet_name = '29.05.2020'
# naming the sheet automaticaly based on the current date
# d = datetime.datetime.today()
# sheet_name = d.strftime('%d.%m.%y')

os.environ["GOOGLE_APPLICATION_CREDENTIALS"]="C:/Users/hamra/PycharmProjects/Sentiment_Analysis_AutoML/key.json"

# =============================================== Writing Results to Excel File ========================================


def write_to_a_new_excel():
    writer = pd.ExcelWriter(myExcel, engine='openpyxl', mode='w')
    text_sentiment.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()


def write_to_existing_excel_new_sheet():
    # writer = pd.ExcelWriter(myExcel, engine='openpyxl', mode='a')
    # writer.book = load_workbook(myExcel)
    text_sentiment.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.save()


def write_to_existing_excel_same_sheet():
    # writer = pd.ExcelWriter(myExcel, engine='openpyxl', mode='a')
    # writer.book = load_workbook(myExcel)
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    reader = pd.read_excel(myExcel, sheet_name=sheet_name)
    text_sentiment.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(reader)+1)
    writer.save()

# write_to_existing_excel_new_sheet()
# write_to_existing_excel_same_sheet()
# write_to_a_new_excel()
# =============================================== Calling the model for prediction =====================================


def get_prediction(content,project_id, model_id):
    prediction_client = automl_v1.PredictionServiceClient()

    name = 'projects/{}/locations/eu/models/{}'.format(project_id, model_id)
    payload = {'text_snippet': {'content': content, 'mime_type': 'text/plain'}}
    params = {}
    request = prediction_client.predict(name, payload, params)
    # request = [request.payload]
    # results = pd.DataFrame(request)
    # results.to_csv('results.csv', mode='a')
    # print(results)
    return request  # waits until request is returned

# ======================== Read only one file from the specified bucket in google and do the analysis ==================


client = storage.Client()
bucket = client.get_bucket('finwedge-news-letter')
# blob = bucket.get_blob('nlpTest.txt') // for getting a specified file in a bucket, we can give the exact name.
# Fore getting a list of files in a bucket : bucket.list_blobs()
# list_blobs(prefix='to only return files that start with this prefix', delimiter='remove a file with a certain suffix
# from the returned list')
blob_list = list(bucket.list_blobs(prefix='2020/May/29.05.2020'))
# to delete the two first blobs which are b'placeholders!!!!
for i in range(1):
    blob_list.pop(0)

for blob in blob_list:
    downloaded_blob = blob.download_as_string()
    #downloaded_blob = str(downloaded_blob)
    #downloaded_blob = str(downloaded_blob)
    #downloaded_blob.encode('unicode_escape').strip()
    #downloaded_blob.encode('utf-8').strip()
    clean_text = re.sub(r'[^a-zA-Z0-9 ,*\u2019-]', u'', downloaded_blob.decode('unicode_escape'), 0, re.UNICODE).encode("utf8")
    clean_text = str(clean_text).replace("b'", "")
    print(clean_text)

    sentiment_results = get_prediction(clean_text, project_id, model_id)
    overall_sentiment = [items for items in sentiment_results.payload]
    overall_metadata = [items for items in sentiment_results.metadata.items()]

    text_sentiment = pd.DataFrame({
        'News Content': clean_text,
        'Sentiment': overall_sentiment,
        'Metadata': overall_metadata
    })
    # checking if the Excel file exists, If NO, it makes a new excel file. If YES, it checks if the sheet name exists,
    # if yes we write the results to the same sheet if not it makes a new sheet and write the info on there.
    if myExcel.is_file():
        # making a excel writer and loading the workbook is needed in both of below functions
        writer = pd.ExcelWriter(myExcel, engine='openpyxl', mode='a')
        writer.book = load_workbook(myExcel)
        if sheet_name in writer.book:
            write_to_existing_excel_same_sheet()
        else:
            write_to_existing_excel_new_sheet()
    else:
        write_to_a_new_excel()

# print(text_sentiment)

# =====================Reading multiple files.txt in a folder on local and do sentiment analysis on each================

"""
for filenames in os.listdir('C:/Users/hamra/PycharmProjects/Sentiment_Analysis_AutoML'):
    if filenames.endswith(".txt"):


        text = open(filenames)
        content = text.read()
        sentiment_results = get_prediction(content, project_id, model_id)
        overall_sentiment = [items for items in sentiment_results.payload]
        overall_metadata = [items for items in sentiment_results.metadata.items()]

        text_sentiment = pd.DataFrame({
            'News Content':content,
            'Sentiment': overall_sentiment,
            'Metadata': overall_metadata
        })
"""